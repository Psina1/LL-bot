from __future__ import annotations

import asyncio
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal
import json
from pathlib import Path
import re
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import text

from app.config import get_settings
from app.db.session import SessionLocal


MOSCOW = ZoneInfo("Europe/Moscow")
SYSTEM_MESSAGE_MODES = {"materials_summary", "materials_podcast_summary"}
BUTTON_CATALOG = [
    ("reply_button", "Задать вопрос", "Главное меню"),
    ("reply_button", "Материалы программы", "Главное меню"),
    ("reply_button", "Домашние задания", "Главное меню"),
    ("reply_button", "Расписание Лиги Лидеров", "Главное меню"),
    ("reply_button", "Настройки уведомлений", "Главное меню"),
    ("inline_button", "Вопрос по программе", "Вопросы"),
    ("inline_button", "Технический вопрос", "Вопросы"),
    ("inline_button", "Другое", "Вопросы"),
    ("inline_button", "Домашние задания: список", "Домашние задания"),
    ("inline_button", "Домашние задания: конкретное ДЗ", "Домашние задания"),
    ("inline_button", "Домашние задания: помощь", "Домашние задания"),
    ("inline_button", "Домашние задания: вопрос по конкретному ДЗ", "Домашние задания"),
    ("inline_button", "Материалы: записи и материалы", "Материалы"),
    ("inline_button", "Материалы: последнее занятие", "Материалы"),
    ("inline_button", "Материалы: выбрать занятие", "Материалы"),
    ("inline_button", "Материалы: все материалы", "Материалы"),
    ("inline_button", "Материалы: блок", "Материалы"),
    ("inline_button", "Материалы: занятие", "Материалы"),
    ("inline_button", "Материалы: материалы занятия", "Материалы"),
    ("inline_button", "Материалы: саммари", "Материалы"),
    ("inline_button", "Материалы: саммари занятия", "Материалы"),
    ("inline_button", "Материалы: скачать оригинал", "Материалы"),
    ("inline_button", "Материалы: подкасты", "Материалы"),
    ("inline_button", "Материалы: подкаст занятия", "Материалы"),
    ("inline_button", "Материалы: открыть подкаст", "Материалы"),
    ("inline_button", "Материалы: видео занятия", "Материалы"),
    ("inline_button", "Материалы: домашка занятия", "Материалы"),
    ("inline_button", "Материалы: открыть саммари сообщением", "Материалы"),
    ("inline_button", "Главное меню", "Навигация"),
    ("reply_button", "Главное меню", "Навигация"),
    ("inline_button", "Старт: выбор времени уведомлений", "Уведомления"),
    ("reply_button", "Уведомления: 09:00", "Уведомления"),
    ("reply_button", "Уведомления: 12:00", "Уведомления"),
    ("reply_button", "Уведомления: 15:00", "Уведомления"),
    ("reply_button", "Уведомления: отключить", "Уведомления"),
    ("inline_button", "Опрос ОС: оценить", "Обратная связь"),
    ("inline_button", "Опрос ОС: не был", "Обратная связь"),
    ("inline_button", "Опрос ОС: напомнить позже", "Обратная связь"),
    ("inline_button", "Опрос ОС: оценка", "Обратная связь"),
]
MODE_LABELS = {
    "free_text": "Свободный вопрос",
    "program_question": "Вопрос по программе",
    "technical_question": "Технический вопрос",
    "other_question": "Другое",
    "training_qa": "Вопрос по обучению",
    "homework_help": "Помощь с домашкой",
    "voice_question": "Голосовой вопрос",
    "bot_voice_capabilities": "Возможности бота",
    "bot_capabilities": "Возможности бота",
    "off_topic": "Вне тематики",
    "project_help": "Помощь с проектом",
    "user_file_qa": "Вопрос по файлу",
    "followup": "Уточняющий вопрос",
}
STOP_WORDS = {
    "когда", "какие", "какой", "которая", "которые", "этого", "чтобы", "можно",
    "нужно", "будет", "было", "есть", "меня", "тебя", "тебе", "себя", "свой",
    "свои", "своего", "просто", "почему", "через", "после", "перед", "если",
    "ответ", "вопрос", "бота", "бот", "программы", "лидеров",
}


def _json_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def _msk(value: datetime | None) -> str:
    if value is None:
        return ""
    if value.tzinfo is None:
        value = value.replace(tzinfo=ZoneInfo("UTC"))
    return value.astimezone(MOSCOW).strftime("%d.%m.%Y, %H:%M")


def _iso(value: datetime | None) -> str | None:
    return value.isoformat() if value else None


def _username(value: str | None) -> str:
    if not value:
        return ""
    return value if value.startswith("@") else f"@{value}"


def _answer_status(answer: str) -> str:
    normalized = answer.lower()
    if "не получилось получить ответ" in normalized or "техническая ошибка" in normalized:
        return "техническая ошибка"
    if "не нашёл точного ответа" in normalized or "не нашел точного ответа" in normalized:
        return "ответ не найден"
    return "ответ получен"


async def _fetch(sql: str, params: dict[str, Any] | None = None) -> list[dict[str, Any]]:
    async with SessionLocal() as session:
        result = await session.execute(text(sql), params or {})
        return [dict(row) for row in result.mappings().all()]


async def build_report() -> dict[str, Any]:
    settings = get_settings()
    admin_ids = set(settings.admin_ids)
    now = datetime.now(MOSCOW)

    allowed = await _fetch("SELECT * FROM allowed_users WHERE is_active IS TRUE ORDER BY id")
    users = await _fetch("SELECT * FROM users ORDER BY id")
    events = await _fetch("SELECT * FROM user_events ORDER BY created_at")
    messages = await _fetch(
        """
        SELECT m.*, u.telegram_id, u.username, u.full_name, u.role::text AS role
        FROM messages m JOIN users u ON u.id=m.user_id
        ORDER BY m.created_at
        """
    )
    notification_settings = await _fetch("SELECT * FROM user_notification_settings")
    deliveries = await _fetch(
        """
        SELECT d.*, u.telegram_id, u.username, u.full_name AS user_name
        FROM notification_deliveries d JOIN users u ON u.id=d.user_id
        ORDER BY d.created_at
        """
    )
    user_files = await _fetch("SELECT user_id, count(1) AS count FROM user_files GROUP BY user_id")
    documents = await _fetch(
        """
        SELECT d.*, u.telegram_id AS owner_telegram_id, count(c.id) AS chunk_count
        FROM documents d
        LEFT JOIN users u ON u.id=d.owner_user_id
        LEFT JOIN chunks c ON c.document_id=d.id
        GROUP BY d.id, u.telegram_id
        ORDER BY d.id
        """
    )
    chunks_count = (await _fetch("SELECT count(1) AS count FROM chunks"))[0]["count"]
    homeworks = await _fetch("SELECT * FROM homeworks ORDER BY id")
    lessons = await _fetch("SELECT * FROM program_lessons ORDER BY sort_order, id")
    media = await _fetch("SELECT * FROM program_media ORDER BY id")
    errors = await _fetch("SELECT * FROM errors ORDER BY created_at DESC")
    message_feedback = await _fetch(
        """
        SELECT f.*, u.telegram_id, u.username
        FROM message_feedback f JOIN users u ON u.id=f.user_id
        ORDER BY f.created_at
        """
    )
    feedback_campaigns = await _fetch(
        """
        SELECT c.*,
               count(DISTINCT r.id) AS recipient_count,
               count(DISTINCT r.id) FILTER (WHERE r.status='completed') AS completed_count,
               count(DISTINCT r.id) FILTER (WHERE r.status='in_progress') AS in_progress_count,
               count(DISTINCT r.id) FILTER (WHERE r.status='pending') AS pending_count,
               count(DISTINCT r.id) FILTER (WHERE r.status='not_attended') AS not_attended_count,
               count(DISTINCT r.id) FILTER (WHERE r.status='declined') AS declined_count,
               round(avg(r.usefulness_score) FILTER (WHERE r.usefulness_score IS NOT NULL), 2) AS usefulness_avg,
               round(avg(r.experts_score) FILTER (WHERE r.experts_score IS NOT NULL), 2) AS experts_avg,
               count(DISTINCT d.response_id) FILTER (
                   WHERE d.status='sent' AND d.delivery_type='initial'
               ) AS initial_sent_count,
               count(DISTINCT d.id) FILTER (WHERE d.status='sent') AS delivery_sent_count,
               count(DISTINCT d.id) FILTER (WHERE d.status='error') AS error_count
        FROM feedback_campaigns c
        LEFT JOIN feedback_responses r ON r.campaign_id=c.id
        LEFT JOIN feedback_deliveries d ON d.campaign_id=c.id
        GROUP BY c.id ORDER BY c.id
        """
    )
    feedback_responses = await _fetch(
        """
        SELECT r.*, c.lesson_title, c.lesson_date, c.is_test,
               u.telegram_id, u.username,
               coalesce(a.full_name, u.full_name, u.username, u.telegram_id::text) AS user_name
        FROM feedback_responses r
        JOIN feedback_campaigns c ON c.id=r.campaign_id
        JOIN users u ON u.id=r.user_id
        LEFT JOIN allowed_users a ON a.telegram_id=u.telegram_id
        ORDER BY r.created_at
        """
    )

    users_by_tg = {int(row["telegram_id"]): row for row in users}
    events_by_tg: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for row in events:
        events_by_tg[int(row["telegram_id"])].append(row)
    messages_by_user: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for row in messages:
        messages_by_user[int(row["user_id"])].append(row)
    settings_by_user = {int(row["user_id"]): row for row in notification_settings}
    deliveries_by_user: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for row in deliveries:
        deliveries_by_user[int(row["user_id"])].append(row)
    files_by_user = {int(row["user_id"]): int(row["count"]) for row in user_files}

    participant_allowed = [
        row for row in allowed
        if row["telegram_id"] is not None and int(row["telegram_id"]) not in admin_ids
    ]
    participants = []
    for allowed_user in participant_allowed:
        telegram_id = int(allowed_user["telegram_id"])
        user = users_by_tg.get(telegram_id)
        user_events = events_by_tg.get(telegram_id, [])
        user_messages = messages_by_user.get(int(user["id"]), []) if user else []
        natural_messages = [m for m in user_messages if m["mode"] not in SYSTEM_MESSAGE_MODES]
        setting = settings_by_user.get(int(user["id"])) if user else None
        user_deliveries = deliveries_by_user.get(int(user["id"]), []) if user else []
        if not user:
            notification_status = "не заходил"
            notification_enabled = False
            notification_time = ""
        elif setting and setting["enabled"]:
            notification_status = "включены"
            notification_enabled = True
            notification_time = setting["notification_time"]
        elif setting:
            notification_status = "отключены"
            notification_enabled = False
            notification_time = setting["notification_time"]
        else:
            notification_status = "не настроены"
            notification_enabled = False
            notification_time = ""
        first_activity = user_events[0]["created_at"] if user_events else None
        last_activity = user_events[-1]["created_at"] if user_events else None
        participants.append(
            {
                "allowedUserId": allowed_user["id"],
                "userId": user["id"] if user else None,
                "telegramId": str(telegram_id),
                "username": _username(allowed_user["username"] or (user["username"] if user else None)),
                "fullName": allowed_user["full_name"] or (user["full_name"] if user else None) or _username(allowed_user["username"]) or str(telegram_id),
                "role": "participant",
                "sourceRoster": "Excel" in (allowed_user["note"] or ""),
                "registered": bool(user),
                "registeredAt": _iso(user["created_at"]) if user else None,
                "firstActivity": _iso(first_activity),
                "lastActivity": _iso(last_activity),
                "eventCount": len(user_events),
                "buttonClickCount": sum(e["event_type"] in {"reply_button", "inline_button"} for e in user_events),
                "commandCount": sum(e["event_type"] == "command" for e in user_events),
                "questionCount": len(natural_messages),
                "uploadedFileCount": files_by_user.get(int(user["id"]), 0) if user else 0,
                "notificationStatus": notification_status,
                "notificationEnabled": notification_enabled,
                "notificationTime": notification_time,
                "deliverySentCount": sum(d["status"] == "sent" for d in user_deliveries),
                "deliveryErrorCount": sum(d["status"] == "error" for d in user_deliveries),
                "hasProjectContext": bool(user and user["project_context"]),
                "registeredAtMsk": _msk(user["created_at"]) if user else "",
                "firstActivityMsk": _msk(first_activity),
                "lastActivityMsk": _msk(last_activity),
            }
        )
    participants.sort(key=lambda row: (not row["registered"], -row["buttonClickCount"], row["fullName"]))

    admins = []
    for telegram_id in sorted(admin_ids):
        user = users_by_tg.get(telegram_id)
        if not user:
            continue
        user_events = events_by_tg.get(telegram_id, [])
        setting = settings_by_user.get(int(user["id"]))
        admins.append(
            {
                "telegramId": str(telegram_id),
                "username": _username(user["username"]),
                "fullName": user["full_name"] or user["username"] or str(telegram_id),
                "registeredAt": _iso(user["created_at"]),
                "lastActivity": _iso(user_events[-1]["created_at"]) if user_events else None,
                "eventCount": len(user_events),
                "questionCount": len(messages_by_user.get(int(user["id"]), [])),
                "notificationStatus": "включены" if setting and setting["enabled"] else "не настроены",
                "notificationTime": setting["notification_time"] if setting else "",
            }
        )

    questions = []
    for row in messages:
        token_usage = row["token_usage"] or {}
        role = "admin" if int(row["telegram_id"]) in admin_ids else "participant"
        questions.append(
            {
                "id": str(row["id"]),
                "userId": str(row["user_id"]),
                "telegramId": str(row["telegram_id"]),
                "userName": row["full_name"] or row["username"] or str(row["telegram_id"]),
                "username": _username(row["username"]),
                "role": role,
                "mode": row["mode"],
                "question": row["question"],
                "answer": row["answer"],
                "answerStatus": _answer_status(row["answer"]),
                "systemGenerated": row["mode"] in SYSTEM_MESSAGE_MODES,
                "promptTokens": int(token_usage.get("prompt_tokens", 0) or 0),
                "completionTokens": int(token_usage.get("completion_tokens", 0) or 0),
                "totalTokens": int(token_usage.get("total_tokens", 0) or 0),
                "sources": row["sources"] or [],
                "createdAt": _iso(row["created_at"]),
                "createdAtMsk": _msk(row["created_at"]),
            }
        )
    participant_questions = [
        row for row in questions
        if row["role"] == "participant" and not row["systemGenerated"]
    ]

    participant_tg_ids = {int(row["telegramId"]) for row in participants}
    participant_events = [e for e in events if int(e["telegram_id"]) in participant_tg_ids]
    admin_events = [e for e in events if int(e["telegram_id"]) in admin_ids]
    participant_button_counts = Counter(
        e["event_name"] for e in participant_events
        if e["event_type"] in {"reply_button", "inline_button"}
    )
    admin_button_counts = Counter(
        e["event_name"] for e in admin_events
        if e["event_type"] in {"reply_button", "inline_button"}
    )
    button_catalog = [
        {
            "eventType": event_type,
            "name": name,
            "section": section,
            "count": participant_button_counts.get(name, 0),
        }
        for event_type, name, section in BUTTON_CATALOG
    ]
    button_catalog.sort(key=lambda row: (-row["count"], row["section"], row["name"]))

    activity_by_date: dict[date, dict[str, Any]] = defaultdict(
        lambda: {"clicks": 0, "commands": 0, "questions": 0, "users": set()}
    )
    for event in participant_events:
        event_date = event["created_at"].astimezone(MOSCOW).date()
        bucket = activity_by_date[event_date]
        if event["event_type"] in {"reply_button", "inline_button"}:
            bucket["clicks"] += 1
        if event["event_type"] == "command":
            bucket["commands"] += 1
        bucket["users"].add(int(event["telegram_id"]))
    for question in participant_questions:
        question_date = datetime.fromisoformat(question["createdAt"]).astimezone(MOSCOW).date()
        activity_by_date[question_date]["questions"] += 1
        activity_by_date[question_date]["users"].add(int(question["telegramId"]))
    period_start = min(
        [row["created_at"].astimezone(MOSCOW).date() for row in events] or [now.date()]
    )
    chart_start = max(period_start, now.date() - timedelta(days=20))
    daily_activity = []
    cursor = chart_start
    while cursor <= now.date():
        bucket = activity_by_date[cursor]
        daily_activity.append(
            {
                "date": cursor.isoformat(),
                "clicks": bucket["clicks"],
                "commands": bucket["commands"],
                "questions": bucket["questions"],
                "activeUsers": len(bucket["users"]),
            }
        )
        cursor += timedelta(days=1)

    mode_counts = Counter(row["mode"] for row in participant_questions)
    question_modes = [
        {"name": name, "count": count, "label": MODE_LABELS.get(name, name)}
        for name, count in mode_counts.most_common()
    ]
    keyword_counts: Counter[str] = Counter()
    for row in participant_questions:
        for word in re.findall(r"[a-zа-яё]{4,}", row["question"].lower()):
            if word not in STOP_WORDS:
                keyword_counts[word] += 1
    top_keywords = [{"name": word, "count": count} for word, count in keyword_counts.most_common(20)]

    status_counts = Counter(row["notificationStatus"] for row in participants)
    notification_status = [
        {"name": "Включены", "count": status_counts["включены"]},
        {"name": "Отключены", "count": status_counts["отключены"]},
        {"name": "Не настроены", "count": status_counts["не настроены"]},
        {"name": "Не заходили", "count": status_counts["не заходил"]},
    ]
    notification_times_counter = Counter(
        row["notificationTime"] for row in participants
        if row["notificationEnabled"] and row["notificationTime"]
    )
    notification_times = [
        {"name": time_value, "count": notification_times_counter.get(time_value, 0)}
        for time_value in ("09:00", "12:00", "15:00")
    ]
    delivery_types = [
        {"name": name, "count": count}
        for name, count in Counter(d["notification_key"] for d in deliveries if d["status"] == "sent").most_common()
    ]

    token_prompt = sum(row["promptTokens"] for row in questions)
    token_completion = sum(row["completionTokens"] for row in questions)
    token_total = sum(row["totalTokens"] for row in questions)
    feedback_completed = sum(
        int(row["completed_count"] or 0) for row in feedback_campaigns
    )
    feedback_not_attended = sum(
        int(row["not_attended_count"] or 0) for row in feedback_campaigns
    )
    feedback_declined = sum(
        int(row["declined_count"] or 0) for row in feedback_campaigns
    )
    registered_participants = sum(row["registered"] for row in participants)
    participants_with_clicks = sum(row["buttonClickCount"] > 0 for row in participants)
    participants_with_questions = sum(row["questionCount"] > 0 for row in participants)
    notifications_enabled = sum(row["notificationEnabled"] for row in participants)
    last_week = now - timedelta(days=7)
    active_last_week = sum(
        bool(row["lastActivity"] and datetime.fromisoformat(row["lastActivity"]).astimezone(MOSCOW) >= last_week)
        for row in participants
    )
    participant_clicks = sum(row["buttonClickCount"] for row in participants)

    kpis = {
        "sourceRosterCount": sum(row["sourceRoster"] for row in participants),
        "allowedParticipants": len(participants),
        "registeredParticipants": registered_participants,
        "neverEntered": len(participants) - registered_participants,
        "participantsWithClicks": participants_with_clicks,
        "participantsWithQuestions": participants_with_questions,
        "naturalQuestions": len(participant_questions),
        "allParticipantMessages": sum(row["role"] == "participant" for row in questions),
        "participantClicks": participant_clicks,
        "notificationsEnabled": notifications_enabled,
        "notificationsNotEnabled": len(participants) - notifications_enabled,
        "activeLast7Days": active_last_week,
        "adminCount": len(admins),
        "adminMessages": sum(row["role"] == "admin" for row in questions),
        "adminClicks": sum(
            event["event_type"] in {"reply_button", "inline_button"} for event in admin_events
        ),
        "deliveriesSent": sum(row["status"] == "sent" for row in deliveries),
        "deliveryErrors": sum(row["status"] == "error" for row in deliveries),
        "tokenPrompt": token_prompt,
        "tokenCompletion": token_completion,
        "tokenTotal": token_total,
        "documentCount": len(documents),
        "readyDocumentCount": sum(str(row["status"]) == "ready" for row in documents),
        "chunkCount": int(chunks_count),
        "homeworkCount": len(homeworks),
        "mediaCount": len(media),
        "errorCount": len(errors),
        "feedbackCampaignCount": len(feedback_campaigns),
        "feedbackResponseCount": len(feedback_responses),
        "feedbackCompletedCount": feedback_completed,
        "feedbackNotAttendedCount": feedback_not_attended,
        "feedbackDeclinedCount": feedback_declined,
    }

    for row in deliveries:
        row["createdAtMsk"] = _msk(row["created_at"])
    for row in events:
        row["createdAtMsk"] = _msk(row["created_at"])
    for row in feedback_campaigns:
        row["createdAtMsk"] = _msk(row["created_at"])
        row["launchAtMsk"] = _msk(row["launch_at"])
    for row in feedback_responses:
        row["createdAtMsk"] = _msk(row["created_at"])
        row["completedAtMsk"] = _msk(row["completed_at"])

    return {
        "meta": {
            "title": "Аналитика использования бота «Помощник Лиги Лидеров»",
            "generatedAt": now.isoformat(),
            "generatedAtMsk": now.strftime("%d.%m.%Y, %H:%M"),
            "periodStart": period_start.isoformat(),
            "periodEnd": now.date().isoformat(),
            "timezone": "Europe/Moscow",
            "notes": [
                "Основные показатели рассчитаны по участникам и не включают активность администраторов.",
                "Статус «не заходил» означает: человек есть в списке доступа, но записи пользователя в БД ещё нет.",
                "Автоматически сформированные запросы саммари и подкаст-выжимок отделены от вопросов, написанных людьми.",
            ],
        },
        "kpis": kpis,
        "funnel": [
            {"name": "Есть доступ", "count": len(participants)},
            {"name": "Зашли в бота", "count": registered_participants},
            {"name": "Нажали кнопку", "count": participants_with_clicks},
            {"name": "Задали вопрос", "count": participants_with_questions},
            {"name": "Включили уведомления", "count": notifications_enabled},
        ],
        "participants": participants,
        "neverEntered": [
            {
                "fullName": row["fullName"],
                "username": row["username"],
                "telegramId": row["telegramId"],
                "sourceRoster": row["sourceRoster"],
            }
            for row in participants if not row["registered"]
        ],
        "admins": admins,
        "questions": questions,
        "participantQuestions": participant_questions,
        "buttonCounts": [{"name": name, "count": count} for name, count in participant_button_counts.most_common()],
        "adminButtonCounts": [{"name": name, "count": count} for name, count in admin_button_counts.most_common()],
        "buttonCatalog": button_catalog,
        "dailyActivity": daily_activity,
        "questionModes": question_modes,
        "topKeywords": top_keywords,
        "notificationStatus": notification_status,
        "notificationTimes": notification_times,
        "deliveries": deliveries,
        "deliveryTypes": delivery_types,
        "documents": documents,
        "homeworks": homeworks,
        "lessons": lessons,
        "media": media,
        "errors": errors,
        "feedback": message_feedback,
        "feedbackCampaigns": feedback_campaigns,
        "feedbackResponses": feedback_responses,
        "rawEvents": events,
    }


def _write_sheet(workbook: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet(title)
    if not rows:
        sheet.append(["Данных пока нет"])
        return
    headers = list(rows[0].keys())
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="07563E")
        cell.alignment = Alignment(vertical="center")
    for row in rows:
        sheet.append(
            [
                json.dumps(row.get(header), ensure_ascii=False)
                if isinstance(row.get(header), (dict, list))
                else _json_value(row.get(header))
                for header in headers
            ]
        )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, header in enumerate(headers, start=1):
        values = [str(header)] + [str(row.get(header, "") or "") for row in rows[:200]]
        width = min(max(max(map(len, values)) + 2, 10), 60)
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_workbook(report: dict[str, Any], output_path: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Сводка"
    summary.append(["Показатель", "Значение"])
    for cell in summary[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="07563E")
    labels = {
        "allowedParticipants": "Участники с доступом",
        "registeredParticipants": "Зашли в бота",
        "neverEntered": "Не заходили",
        "participantsWithQuestions": "Задавали вопросы",
        "naturalQuestions": "Вопросов участников",
        "participantClicks": "Нажатий участников",
        "notificationsEnabled": "Уведомления включены",
        "documentCount": "Документы",
        "chunkCount": "Чанки",
        "homeworkCount": "Домашние задания",
        "mediaCount": "Медиа",
        "feedbackCampaignCount": "Кампании обратной связи",
        "feedbackResponseCount": "Получатели опросов",
        "feedbackCompletedCount": "Завершили опросы",
        "feedbackNotAttendedCount": "Не были на занятии",
        "feedbackDeclinedCount": "Отказались от опроса",
        "errorCount": "Ошибки",
    }
    for key, label in labels.items():
        summary.append([label, report["kpis"][key]])
    summary.column_dimensions["A"].width = 38
    summary.column_dimensions["B"].width = 18
    _write_sheet(workbook, "Участники", report["participants"])
    _write_sheet(workbook, "Вопросы", report["participantQuestions"])
    _write_sheet(workbook, "Кнопки", report["buttonCatalog"])
    _write_sheet(workbook, "Уведомления", report["deliveries"])
    _write_sheet(workbook, "Кампании ОС", report["feedbackCampaigns"])
    _write_sheet(workbook, "Ответы ОС", report["feedbackResponses"])
    _write_sheet(workbook, "Документы", report["documents"])
    _write_sheet(workbook, "Домашние задания", report["homeworks"])
    _write_sheet(workbook, "Медиа", report["media"])
    _write_sheet(workbook, "Ошибки", report["errors"])
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


async def main() -> None:
    output_dir = Path("/tmp/ll_bot_analytics")
    output_dir.mkdir(parents=True, exist_ok=True)
    report = await build_report()
    json_path = output_dir / "report.json"
    xlsx_path = output_dir / "LL_bot_analytics.xlsx"
    json_path.write_text(
        json.dumps(report, ensure_ascii=False, default=_json_value),
        encoding="utf-8",
    )
    write_workbook(report, xlsx_path)
    print(
        json.dumps(
            {
                "json": str(json_path),
                "xlsx": str(xlsx_path),
                "participants": report["kpis"]["allowedParticipants"],
                "registered": report["kpis"]["registeredParticipants"],
                "questions": report["kpis"]["naturalQuestions"],
                "campaigns": report["kpis"]["feedbackCampaignCount"],
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    asyncio.run(main())
