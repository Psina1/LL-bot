from __future__ import annotations

from datetime import date, datetime, timedelta
from io import BytesIO
import logging

from aiogram import F, Router
from aiogram.exceptions import TelegramBadRequest
from aiogram.filters import BaseFilter
from aiogram.fsm.context import FSMContext
from aiogram.types import BufferedInputFile, CallbackQuery, Message
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.bot.keyboards.lesson_feedback import (
    admin_feedback_campaign_keyboard,
    admin_feedback_launch_keyboard,
    admin_feedback_lessons_keyboard,
    admin_feedback_menu_keyboard,
    feedback_score_keyboard,
)
from app.bot.keyboards.reply import admin_menu_keyboard, all_reply_button_labels
from app.bot.states.forms import AdminFlow, UserFlow
from app.db.feedback_repositories import (
    FeedbackCampaignRepository,
    FeedbackResponseRepository,
)
from app.db.repositories import ErrorRepository, ProgramLessonRepository, UserRepository
from app.db.session import SessionLocal
from app.services.container import AppContainer

logger = logging.getLogger(__name__)

DEFAULT_SECOND_SCORE_QUESTION = "Насколько ты доволен выступлением эксперта(-ов)?"
GROUP_WORK_SECOND_SCORE_QUESTION = "Насколько полезной для тебя была работа в группе?"
GROUP_WORK_FEEDBACK_DATES = {date(2026, 6, 16)}
DEFAULT_USEFULNESS_QUESTION = "Насколько полезным для тебя было занятие?"
DEFAULT_VALUABLE_QUESTION = "Что было ценным?"
DEFAULT_IMPROVEMENT_QUESTION = "Что стоит улучшить?"
FEEDBACK_QUESTION_FIELDS = (
    "usefulness_question",
    "experts_question",
    "valuable_question",
    "improvement_question",
)


def build_lesson_display_numbers(lessons) -> dict[str, int]:
    """Number actual lessons in the order they appear in the full program."""
    numbered_lessons = [lesson for lesson in lessons if lesson.lesson_number is not None]
    return {
        lesson.lesson_key: index
        for index, lesson in enumerate(numbered_lessons, start=1)
    }


def second_score_question_text(lesson_date) -> str:
    if lesson_date in GROUP_WORK_FEEDBACK_DATES:
        return GROUP_WORK_SECOND_SCORE_QUESTION
    return DEFAULT_SECOND_SCORE_QUESTION


def default_feedback_questions(lesson_date) -> dict[str, str]:
    return {
        "usefulness_question": DEFAULT_USEFULNESS_QUESTION,
        "experts_question": second_score_question_text(lesson_date),
        "valuable_question": DEFAULT_VALUABLE_QUESTION,
        "improvement_question": DEFAULT_IMPROVEMENT_QUESTION,
    }


def campaign_feedback_questions(campaign) -> dict[str, str]:
    defaults = default_feedback_questions(campaign.lesson_date if campaign else None)
    if campaign is None:
        return defaults
    return {
        field: (getattr(campaign, field, None) or defaults[field]).strip()
        for field in FEEDBACK_QUESTION_FIELDS
    }


def clean_feedback_question(text: str) -> str:
    return " ".join(text.strip().split())


class ActiveFeedbackOpenAnswerFilter(BaseFilter):
    def __init__(self) -> None:
        self.reply_button_labels = all_reply_button_labels()

    async def __call__(self, message: Message, state: FSMContext | None = None) -> dict | bool:
        if state is not None:
            current_state = await state.get_state()
            if current_state in {
                UserFlow.waiting_for_bot_feedback_score.state,
                UserFlow.waiting_for_bot_feedback_useful.state,
                UserFlow.waiting_for_bot_feedback_improve.state,
                UserFlow.waiting_for_bot_feedback_missing.state,
            }:
                return False
        if message.text:
            text = message.text.strip()
            if text.startswith("/") or text in self.reply_button_labels:
                return False
        async with SessionLocal() as session:
            response = await FeedbackResponseRepository.get_open_answer_for_telegram_user(
                session,
                message.from_user.id,
            )
        if response is None:
            return False
        return {"lesson_feedback_response_id": response.id}


def build_lesson_feedback_router(container: AppContainer) -> Router:
    router = Router(name="lesson_feedback")
    open_answer_filter = ActiveFeedbackOpenAnswerFilter()

    def is_admin(telegram_id: int) -> bool:
        return telegram_id in container.settings.admin_ids

    async def upsert_user(telegram_user):
        async with SessionLocal() as session:
            return await UserRepository.upsert_telegram_user(
                session=session,
                telegram_id=telegram_user.id,
                username=telegram_user.username,
                first_name=telegram_user.first_name,
                last_name=telegram_user.last_name,
                is_admin=is_admin(telegram_user.id),
            )

    async def get_response_for_callback(callback: CallbackQuery, response_id: int):
        user = await upsert_user(callback.from_user)
        async with SessionLocal() as session:
            response = await FeedbackResponseRepository.get_for_user(
                session,
                response_id,
                user.id,
            )
            campaign = (
                await FeedbackCampaignRepository.get(session, response.campaign_id)
                if response is not None
                else None
            )
        if response is None or campaign is None:
            await callback.answer("Этот опрос не найден.", show_alert=True)
            return None, None, None
        if campaign.status != "active":
            await callback.answer("Этот опрос уже закрыт.", show_alert=True)
            return None, None, None
        return user, response, campaign

    async def lesson_display_numbers() -> dict[str, int]:
        async with SessionLocal() as session:
            lessons = await ProgramLessonRepository.list_active(session)
        return build_lesson_display_numbers(lessons)

    def state_feedback_questions(data: dict, lesson_date) -> dict[str, str]:
        defaults = default_feedback_questions(lesson_date)
        return {
            field: clean_feedback_question(data.get(field) or defaults[field])
            for field in FEEDBACK_QUESTION_FIELDS
        }

    def feedback_questions_text(questions: dict[str, str]) -> str:
        return (
            "Тексты вопросов:\n"
            f"1. {questions['usefulness_question']}\n"
            f"2. {questions['experts_question']}\n"
            f"3. {questions['valuable_question']}\n"
            f"4. {questions['improvement_question']}"
        )

    async def show_feedback_launch_card(message: Message, lesson, state: FSMContext) -> None:
        display_number = (await lesson_display_numbers()).get(lesson.lesson_key)
        survey_label = f"Опрос №{display_number}" if display_number else "Опрос обратной связи"
        date_text = lesson.date_start.strftime("%d.%m.%Y") if lesson.date_start else "без даты"
        experts = lesson.speaker or "не указаны"
        state_data = await state.get_data()
        questions = state_feedback_questions(state_data, lesson.date_start)
        await message.answer(
            f"{survey_label}:\n"
            f"{lesson.lesson_title}\n"
            f"Дата: {date_text}\n"
            f"Эксперты: {experts}\n\n"
            f"{feedback_questions_text(questions)}\n\n"
            "Сначала можно отправить безопасный тест только администраторам.\n\n"
            "Если нажать «Запустить через 10 минут», в течение этих 10 минут "
            "рассылку можно отменить кнопкой «Закрыть опрос».",
            reply_markup=admin_feedback_launch_keyboard(lesson.lesson_key),
            parse_mode=None,
        )

    async def ask_feedback_question_step(
        message: Message,
        state: FSMContext,
        step_number: int,
    ) -> None:
        data = await state.get_data()
        lesson_date_raw = data.get("feedback_lesson_date")
        lesson_date = date.fromisoformat(lesson_date_raw) if lesson_date_raw else None
        questions = state_feedback_questions(data, lesson_date)
        labels = {
            1: "Вопрос 1 из 4: оценка полезности занятия",
            2: "Вопрос 2 из 4: оценка эксперта(-ов) или формата",
            3: "Вопрос 3 из 4: открытый вопрос про ценность",
            4: "Вопрос 4 из 4: открытый вопрос про улучшения",
        }
        fields = {
            1: "usefulness_question",
            2: "experts_question",
            3: "valuable_question",
            4: "improvement_question",
        }
        await message.answer(
            f"{labels[step_number]}.\n\n"
            f"Сейчас:\n{questions[fields[step_number]]}\n\n"
            "Напиши новый текст одним сообщением.\n"
            "Если текст менять не нужно, отправь `-`.",
            parse_mode=None,
        )

    async def show_feedback_menu(message: Message) -> None:
        await message.answer(
            "Здесь можно запустить опрос после занятия и посмотреть ответы участников.\n\n"
            "Перед общей рассылкой можно отправить безопасный тест администраторам. "
            "Без явного запуска опрос участникам не уйдёт.",
            reply_markup=admin_feedback_menu_keyboard(),
            parse_mode=None,
        )

    async def show_feedback_lessons(message: Message) -> None:
        async with SessionLocal() as session:
            lessons = await ProgramLessonRepository.list_active(session)
        display_numbers = await lesson_display_numbers()
        today = datetime.now(container.feedback_service.timezone).date()
        past_lessons = [
            lesson
            for lesson in lessons
            if lesson.date_start is not None and lesson.date_start <= today
        ]
        past_lessons.sort(key=lambda lesson: (lesson.date_start, lesson.sort_order), reverse=True)
        await message.answer(
            "Выбери занятие, по которому нужно собрать обратную связь.",
            reply_markup=admin_feedback_lessons_keyboard(past_lessons[:12], display_numbers),
            parse_mode=None,
        )

    async def show_admin_campaigns(message: Message) -> None:
        display_numbers = await lesson_display_numbers()
        async with SessionLocal() as session:
            campaigns = await FeedbackCampaignRepository.list_recent(session, limit=10)
            campaign_data = []
            for campaign in campaigns:
                totals = await FeedbackResponseRepository.campaign_totals(session, campaign.id)
                campaign_data.append((campaign, totals))

        if not campaign_data:
            await message.answer(
                "Опросы обратной связи пока не запускались.",
                reply_markup=admin_feedback_menu_keyboard(),
            )
            return

        await message.answer(
            "Активные и последние опросы:",
            reply_markup=admin_feedback_menu_keyboard(),
        )
        for campaign, totals in campaign_data:
            display_number = display_numbers.get(campaign.lesson_key)
            campaign_label = (
                f"Опрос №{display_number}"
                if display_number
                else "Опрос обратной связи"
            )
            date_text = campaign.lesson_date.strftime("%d.%m.%Y") if campaign.lesson_date else "без даты"
            completed = totals.get("completed", 0)
            total = sum(totals.values())
            status_label = {
                "scheduled": "запланирован",
                "active": "активен",
                "closed": "закрыт",
            }.get(campaign.status, campaign.status)
            test_label = " · тест" if campaign.is_test else ""
            await message.answer(
                f"{campaign_label}{test_label}\n"
                f"{campaign.lesson_title}\n"
                f"Дата: {date_text}\n"
                f"Статус: {status_label}\n"
                f"Заполнили: {completed} из {total}",
                reply_markup=admin_feedback_campaign_keyboard(
                    campaign.id,
                    can_close=campaign.status in {"scheduled", "active"},
                ),
                parse_mode=None,
            )

    @router.message(F.text.in_(["Обратная связь", "Админ: обратная связь"]))
    async def admin_feedback_menu(message: Message, state: FSMContext) -> None:
        if not is_admin(message.from_user.id):
            return
        await state.clear()
        await upsert_user(message.from_user)
        await show_feedback_menu(message)

    @router.callback_query(F.data == "admin_lesson_feedback:menu")
    async def admin_feedback_menu_callback(callback: CallbackQuery) -> None:
        if not is_admin(callback.from_user.id):
            await callback.answer("Доступно только администраторам.", show_alert=True)
            return
        await callback.answer()
        if callback.message:
            await show_feedback_menu(callback.message)

    @router.callback_query(F.data == "admin_lesson_feedback:new")
    async def admin_feedback_new(callback: CallbackQuery) -> None:
        if not is_admin(callback.from_user.id):
            await callback.answer("Доступно только администраторам.", show_alert=True)
            return
        await callback.answer()
        if callback.message:
            await show_feedback_lessons(callback.message)

    @router.callback_query(F.data == "admin_lesson_feedback:list")
    async def admin_feedback_list(callback: CallbackQuery) -> None:
        if not is_admin(callback.from_user.id):
            await callback.answer("Доступно только администраторам.", show_alert=True)
            return
        await callback.answer()
        if callback.message:
            await show_admin_campaigns(callback.message)

    @router.callback_query(F.data == "admin_lesson_feedback:cancel")
    async def admin_feedback_cancel(callback: CallbackQuery) -> None:
        if not is_admin(callback.from_user.id):
            await callback.answer("Доступно только администраторам.", show_alert=True)
            return
        await callback.answer("Запуск отменён.")
        if callback.message:
            await callback.message.answer(
                "Запуск опроса отменён. Участникам ничего не отправлено.",
                reply_markup=admin_menu_keyboard(),
                parse_mode=None,
            )

    @router.callback_query(F.data.startswith("admin_lesson_feedback:lesson:"))
    async def admin_feedback_lesson(callback: CallbackQuery, state: FSMContext) -> None:
        if not is_admin(callback.from_user.id):
            await callback.answer("Доступно только администраторам.", show_alert=True)
            return
        lesson_key = (callback.data or "").split(":", 2)[2]
        async with SessionLocal() as session:
            lesson = await ProgramLessonRepository.get_by_key(session, lesson_key)
        if lesson is None:
            await callback.answer("Занятие не найдено.", show_alert=True)
            return
        await callback.answer()
        if callback.message:
            await state.clear()
            await state.update_data(
                feedback_lesson_key=lesson.lesson_key,
                feedback_lesson_date=lesson.date_start.isoformat() if lesson.date_start else None,
                **default_feedback_questions(lesson.date_start),
            )
            await show_feedback_launch_card(callback.message, lesson, state)

    @router.callback_query(F.data.startswith("admin_lesson_feedback:edit_questions:"))
    async def admin_feedback_edit_questions(callback: CallbackQuery, state: FSMContext) -> None:
        if not is_admin(callback.from_user.id):
            await callback.answer("Доступно только администраторам.", show_alert=True)
            return
        lesson_key = (callback.data or "").split(":", 2)[2]
        async with SessionLocal() as session:
            lesson = await ProgramLessonRepository.get_by_key(session, lesson_key)
        if lesson is None:
            await callback.answer("Занятие не найдено.", show_alert=True)
            return
        data = await state.get_data()
        if data.get("feedback_lesson_key") != lesson.lesson_key:
            await state.clear()
            await state.update_data(
                feedback_lesson_key=lesson.lesson_key,
                feedback_lesson_date=lesson.date_start.isoformat() if lesson.date_start else None,
                **default_feedback_questions(lesson.date_start),
            )
        await state.set_state(AdminFlow.waiting_for_feedback_question_1)
        await callback.answer("Запускаю мастер вопросов.")
        if callback.message:
            await ask_feedback_question_step(callback.message, state, 1)

    @router.callback_query(
        F.data.startswith("admin_lesson_feedback:test:")
        | F.data.startswith("admin_lesson_feedback:schedule:")
    )
    async def admin_feedback_launch(callback: CallbackQuery, state: FSMContext) -> None:
        if not is_admin(callback.from_user.id):
            await callback.answer("Доступно только администраторам.", show_alert=True)
            return
        parts = (callback.data or "").split(":", 2)
        action = parts[1]
        lesson_key = parts[2]
        admin_user = await upsert_user(callback.from_user)
        async with SessionLocal() as session:
            lesson = await ProgramLessonRepository.get_by_key(session, lesson_key)
        if lesson is None:
            await callback.answer("Занятие не найдено.", show_alert=True)
            return

        is_test = action == "test"
        delay = timedelta(seconds=2) if is_test else timedelta(minutes=10)
        launch_at = datetime.now(container.feedback_service.timezone) + delay
        state_data = await state.get_data()
        questions = (
            state_feedback_questions(state_data, lesson.date_start)
            if state_data.get("feedback_lesson_key") == lesson.lesson_key
            else default_feedback_questions(lesson.date_start)
        )
        async with SessionLocal() as session:
            campaign = await FeedbackCampaignRepository.create(
                session=session,
                lesson=lesson,
                created_by_user_id=admin_user.id,
                launch_at=launch_at,
                is_test=is_test,
                usefulness_question=questions["usefulness_question"],
                experts_question=questions["experts_question"],
                valuable_question=questions["valuable_question"],
                improvement_question=questions["improvement_question"],
            )

        display_number = (await lesson_display_numbers()).get(lesson.lesson_key)
        survey_label = f"Опрос №{display_number}" if display_number else "Опрос обратной связи"
        if is_test:
            answer = f"{survey_label}: тест отправится администраторам через несколько секунд."
        else:
            answer = (
                f"{survey_label} запланирован через 10 минут.\n"
                "До запуска его можно закрыть в списке опросов."
            )
        await callback.answer("Опрос создан.")
        if callback.message:
            await callback.message.answer(
                answer,
                reply_markup=admin_feedback_campaign_keyboard(campaign.id, can_close=True),
                parse_mode=None,
            )

    async def save_feedback_question_answer(
        message: Message,
        state: FSMContext,
        field: str,
        next_state,
        next_step: int | None,
    ) -> None:
        if not is_admin(message.from_user.id):
            await state.clear()
            return
        text = (message.text or "").strip()
        if text == "Админ: меню":
            await state.clear()
            await message.answer("Редактирование вопросов отменено.", reply_markup=admin_menu_keyboard())
            return
        if text != "-":
            cleaned = clean_feedback_question(text)
            if len(cleaned) < 5:
                await message.answer("Слишком коротко. Напиши вопрос подробнее или отправь `-`, чтобы оставить текущий текст.", parse_mode=None)
                return
            if len(cleaned) > 500:
                await message.answer("Вопрос слишком длинный. Давай до 500 символов, чтобы участникам было легко читать.")
                return
            await state.update_data(**{field: cleaned})
        if next_state is not None and next_step is not None:
            await state.set_state(next_state)
            await ask_feedback_question_step(message, state, next_step)
            return

        await state.set_state(None)
        data = await state.get_data()
        lesson_key = data.get("feedback_lesson_key")
        async with SessionLocal() as session:
            lesson = await ProgramLessonRepository.get_by_key(session, lesson_key) if lesson_key else None
        if lesson is None:
            await message.answer("Не нашёл занятие для опроса. Начни запуск опроса заново.", reply_markup=admin_feedback_menu_keyboard())
            return
        await message.answer("Готово. Вопросы обновлены в черновике опроса.", parse_mode=None)
        await show_feedback_launch_card(message, lesson, state)

    @router.message(AdminFlow.waiting_for_feedback_question_1, F.text)
    async def admin_feedback_question_1(message: Message, state: FSMContext) -> None:
        await save_feedback_question_answer(
            message,
            state,
            "usefulness_question",
            AdminFlow.waiting_for_feedback_question_2,
            2,
        )

    @router.message(AdminFlow.waiting_for_feedback_question_2, F.text)
    async def admin_feedback_question_2(message: Message, state: FSMContext) -> None:
        await save_feedback_question_answer(
            message,
            state,
            "experts_question",
            AdminFlow.waiting_for_feedback_question_3,
            3,
        )

    @router.message(AdminFlow.waiting_for_feedback_question_3, F.text)
    async def admin_feedback_question_3(message: Message, state: FSMContext) -> None:
        await save_feedback_question_answer(
            message,
            state,
            "valuable_question",
            AdminFlow.waiting_for_feedback_question_4,
            4,
        )

    @router.message(AdminFlow.waiting_for_feedback_question_4, F.text)
    async def admin_feedback_question_4(message: Message, state: FSMContext) -> None:
        await save_feedback_question_answer(
            message,
            state,
            "improvement_question",
            None,
            None,
        )

    @router.callback_query(F.data.startswith("admin_lesson_feedback:close:"))
    async def admin_feedback_close(callback: CallbackQuery) -> None:
        if not is_admin(callback.from_user.id):
            await callback.answer("Доступно только администраторам.", show_alert=True)
            return
        try:
            campaign_id = int((callback.data or "").rsplit(":", 1)[1])
        except ValueError:
            await callback.answer("Опрос не найден.", show_alert=True)
            return
        async with SessionLocal() as session:
            await FeedbackCampaignRepository.close(
                session,
                campaign_id,
                datetime.now(container.feedback_service.timezone),
            )
        await callback.answer("Опрос закрыт.")
        if callback.message:
            await callback.message.answer(
                "Опрос закрыт. Новые напоминания по нему отправляться не будут.",
                reply_markup=admin_feedback_menu_keyboard(),
            )

    @router.callback_query(F.data.startswith("admin_lesson_feedback:export:"))
    async def admin_feedback_export(callback: CallbackQuery) -> None:
        if not is_admin(callback.from_user.id):
            await callback.answer("Доступно только администраторам.", show_alert=True)
            return
        try:
            campaign_id = int((callback.data or "").rsplit(":", 1)[1])
        except ValueError:
            await callback.answer("Опрос не найден.", show_alert=True)
            return
        async with SessionLocal() as session:
            rows = await FeedbackResponseRepository.list_export_rows(session, campaign_id)
            campaign = await FeedbackCampaignRepository.get(session, campaign_id)
        if campaign is None:
            await callback.answer("Опрос не найден.", show_alert=True)
            return

        await callback.answer("Готовлю Excel.")
        display_number = (await lesson_display_numbers()).get(campaign.lesson_key)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Ответы"
        questions = campaign_feedback_questions(campaign)
        headers = [
            "Занятие",
            "Дата занятия",
            "Эксперты",
            "ФИО",
            "Telegram",
            "Telegram ID",
            "Статус",
            f"{questions['usefulness_question']} 1-5",
            f"{questions['experts_question']} 1-5",
            questions["valuable_question"],
            "Формат ответа 1",
            questions["improvement_question"],
            "Формат ответа 2",
            "Напоминаний",
            "Дата ответа",
        ]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2563EB")

        status_labels = {
            "pending": "не ответил",
            "in_progress": "заполняет",
            "completed": "завершил",
            "not_attended": "не был на занятии",
            "declined": "не готов оценить",
        }
        for response, row_campaign, user, allowed_user in rows:
            full_name = (
                (allowed_user.full_name if allowed_user else None)
                or user.full_name
                or user.username
                or str(user.telegram_id)
            )
            username = (
                (allowed_user.username if allowed_user else None)
                or user.username
                or ""
            )
            if username and not username.startswith("@"):
                username = f"@{username}"
            sheet.append(
                [
                    row_campaign.lesson_title,
                    row_campaign.lesson_date.strftime("%d.%m.%Y") if row_campaign.lesson_date else "",
                    row_campaign.experts or "",
                    full_name,
                    username,
                    user.telegram_id,
                    status_labels.get(response.status, response.status),
                    response.usefulness_score,
                    response.experts_score,
                    response.valuable_answer or "",
                    response.valuable_input_type or "",
                    response.improvement_answer or "",
                    response.improvement_input_type or "",
                    response.reminder_count,
                    response.completed_at.isoformat() if response.completed_at else "",
                ]
            )
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        widths = [34, 14, 24, 28, 22, 15, 20, 16, 16, 48, 16, 48, 16, 14, 24]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + index)].width = width

        summary = workbook.create_sheet("Сводка")
        totals: dict[str, int] = {}
        completed_scores = []
        experts_scores = []
        for response, _, _, _ in rows:
            totals[response.status] = totals.get(response.status, 0) + 1
            if response.usefulness_score is not None:
                completed_scores.append(response.usefulness_score)
            if response.experts_score is not None:
                experts_scores.append(response.experts_score)
        summary.append(["Показатель", "Значение"])
        summary.append(["Занятие", campaign.lesson_title])
        summary.append(["Всего получателей", len(rows)])
        summary.append(["Завершили", totals.get("completed", 0)])
        summary.append(["Не были на занятии", totals.get("not_attended", 0)])
        summary.append(["Не готовы оценить", totals.get("declined", 0)])
        summary.append(["Не завершили", totals.get("pending", 0) + totals.get("in_progress", 0)])
        summary.append(
            ["Средняя полезность", round(sum(completed_scores) / len(completed_scores), 2) if completed_scores else ""]
        )
        summary.append(
            [
                f"Средняя оценка: {questions['experts_question']}",
                round(sum(experts_scores) / len(experts_scores), 2) if experts_scores else "",
            ]
        )
        for cell in summary[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2563EB")
        summary.column_dimensions["A"].width = 30
        summary.column_dimensions["B"].width = 70

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        if callback.message:
            await callback.message.answer_document(
                BufferedInputFile(
                    output.getvalue(),
                    filename=(
                        f"lesson_feedback_lesson_{display_number}.xlsx"
                        if display_number
                        else f"lesson_feedback_{campaign_id}.xlsx"
                    ),
                ),
                caption=f"Обратная связь: {campaign.lesson_title}",
                reply_markup=admin_menu_keyboard(),
            )

    @router.callback_query(F.data.startswith("lesson_feedback:start:"))
    async def feedback_start(callback: CallbackQuery) -> None:
        response_id = int((callback.data or "").rsplit(":", 1)[1])
        _, response, campaign = await get_response_for_callback(callback, response_id)
        if response is None:
            return
        if response.status in {"completed", "not_attended", "declined"}:
            await callback.answer("Ответ по этому занятию уже сохранён.", show_alert=True)
            return
        async with SessionLocal() as session:
            await FeedbackResponseRepository.start(
                session,
                response_id,
                datetime.now(container.feedback_service.timezone),
            )
        await callback.answer()
        if callback.message:
            questions = campaign_feedback_questions(campaign)
            await callback.message.answer(
                questions["usefulness_question"],
                reply_markup=feedback_score_keyboard(response_id, "usefulness"),
            )

    @router.callback_query(F.data.startswith("lesson_feedback:not_attended:"))
    async def feedback_not_attended(callback: CallbackQuery) -> None:
        response_id = int((callback.data or "").rsplit(":", 1)[1])
        _, response, campaign = await get_response_for_callback(callback, response_id)
        if response is None:
            return
        async with SessionLocal() as session:
            await FeedbackResponseRepository.set_terminal_status(
                session,
                response_id,
                "not_attended",
                datetime.now(container.feedback_service.timezone),
            )
        await callback.answer("Спасибо, записал.")
        if callback.message:
            await callback.message.answer("Спасибо, записал, что тебя не было на занятии.")

    @router.callback_query(F.data.startswith("lesson_feedback:decline:"))
    async def feedback_decline(callback: CallbackQuery) -> None:
        response_id = int((callback.data or "").rsplit(":", 1)[1])
        _, response, campaign = await get_response_for_callback(callback, response_id)
        if response is None:
            return
        async with SessionLocal() as session:
            await FeedbackResponseRepository.set_terminal_status(
                session,
                response_id,
                "declined",
                datetime.now(container.feedback_service.timezone),
            )
        await callback.answer("Ответ сохранён.")
        if callback.message:
            await callback.message.answer("Хорошо, больше не буду напоминать об этом опросе.")

    @router.callback_query(F.data.startswith("lesson_feedback:later:"))
    async def feedback_later(callback: CallbackQuery) -> None:
        response_id = int((callback.data or "").rsplit(":", 1)[1])
        user, response, _ = await get_response_for_callback(callback, response_id)
        if response is None or user is None:
            return
        async with SessionLocal() as session:
            from app.db.repositories import UserNotificationSettingRepository

            setting = await UserNotificationSettingRepository.get_for_user(session, user.id)
            next_reminder_at = container.feedback_service.next_day_at(
                datetime.now(container.feedback_service.timezone),
                setting.notification_time if setting and setting.enabled else "12:00",
            )
            await FeedbackResponseRepository.postpone(
                session,
                response_id,
                next_reminder_at,
            )
        await callback.answer("Напомню позже.")
        if callback.message:
            await callback.message.answer(
                f"Хорошо, напомню завтра в {next_reminder_at.strftime('%H:%M')} по московскому времени."
            )

    @router.callback_query(F.data.startswith("lesson_feedback:score:"))
    async def feedback_score(callback: CallbackQuery) -> None:
        parts = (callback.data or "").split(":")
        if len(parts) != 5:
            await callback.answer("Не понял оценку.", show_alert=True)
            return
        score_type = parts[2]
        response_id = int(parts[3])
        score = int(parts[4])
        if score not in range(1, 6):
            await callback.answer("Оценка должна быть от 1 до 5.", show_alert=True)
            return
        _, response, campaign = await get_response_for_callback(callback, response_id)
        if response is None:
            return
        async with SessionLocal() as session:
            if score_type == "usefulness":
                await FeedbackResponseRepository.set_usefulness_score(
                    session,
                    response_id,
                    score,
                )
            elif score_type == "experts":
                await FeedbackResponseRepository.set_experts_score(
                    session,
                    response_id,
                    score,
                )
            else:
                await callback.answer("Не понял вопрос.", show_alert=True)
                return
        await callback.answer(f"Записал: {score}")
        if not callback.message:
            return
        questions = campaign_feedback_questions(campaign)
        if score_type == "usefulness":
            await callback.message.answer(
                questions["experts_question"],
                reply_markup=feedback_score_keyboard(response_id, "experts"),
            )
        else:
            await callback.message.answer(
                f"{questions['valuable_question']}\n\n"
                "Ответь текстом или отправь голосовое сообщение.",
                parse_mode=None,
            )

    async def save_open_answer(
        message: Message,
        response_id: int,
        answer: str,
        input_type: str,
        voice_file_id: str | None = None,
        transcription_model: str | None = None,
    ) -> None:
        async with SessionLocal() as session:
            response = await FeedbackResponseRepository.get(session, response_id)
            if response is None:
                return
            current_step = response.current_step
            campaign = await FeedbackCampaignRepository.get(session, response.campaign_id)
            questions = campaign_feedback_questions(campaign)
            if current_step == "valuable":
                await FeedbackResponseRepository.set_valuable_answer(
                    session=session,
                    response_id=response_id,
                    answer=answer,
                    input_type=input_type,
                    voice_file_id=voice_file_id,
                    transcription_model=transcription_model,
                )
            elif current_step == "improvement":
                await FeedbackResponseRepository.set_improvement_answer(
                    session=session,
                    response_id=response_id,
                    answer=answer,
                    input_type=input_type,
                    voice_file_id=voice_file_id,
                    transcription_model=transcription_model,
                    now=datetime.now(container.feedback_service.timezone),
                )
            else:
                return

        if current_step == "valuable":
            await message.answer(
                f"{questions['improvement_question']}\n\n"
                "Можно ответить текстом или голосом.",
                parse_mode=None,
            )
        else:
            await message.answer(
                "Спасибо за обратную связь ❤️\n\n"
                "Мы действительно читаем ответы участников и используем их "
                "при проектировании следующих занятий программы.",
                parse_mode=None,
            )

    @router.message(open_answer_filter, F.text)
    async def feedback_open_text(
        message: Message,
        lesson_feedback_response_id: int,
    ) -> None:
        answer = (message.text or "").strip()
        if len(answer) < 2:
            await message.answer("Напиши, пожалуйста, чуть подробнее.")
            return
        await save_open_answer(
            message,
            lesson_feedback_response_id,
            answer,
            input_type="text",
        )

    @router.message(open_answer_filter, F.voice)
    async def feedback_open_voice(
        message: Message,
        lesson_feedback_response_id: int,
    ) -> None:
        voice = message.voice
        if voice.duration > container.settings.max_voice_duration_seconds:
            await message.answer(
                f"Голосовое слишком длинное. Максимум "
                f"{container.settings.max_voice_duration_seconds // 60} минуты."
            )
            return
        if voice.file_size and voice.file_size > container.settings.max_voice_size_bytes:
            await message.answer(
                f"Голосовое слишком большое. Максимум {container.settings.max_voice_size_mb} МБ."
            )
            return
        if not container.transcription_service.enabled:
            await message.answer(
                "Расшифровка голосовых пока не подключена. Ответь, пожалуйста, текстом."
            )
            return

        processing = await message.answer("Расшифровываю голосовое сообщение...")
        try:
            buffer = BytesIO()
            await message.bot.download(voice.file_id, destination=buffer)
            result = await container.transcription_service.transcribe(
                buffer.getvalue(),
                filename="voice.ogg",
                mime_type=voice.mime_type or "audio/ogg",
            )
            await save_open_answer(
                message,
                lesson_feedback_response_id,
                result.text,
                input_type="voice",
                voice_file_id=voice.file_id,
                transcription_model=result.model,
            )
        except Exception as exc:
            logger.exception("lesson_feedback_voice_transcription_failed")
            user = await upsert_user(message.from_user)
            async with SessionLocal() as session:
                await ErrorRepository.create(
                    session,
                    context="lesson_feedback_voice_transcription",
                    error_text=str(exc),
                    user_id=user.id,
                )
            await message.answer(
                "Не получилось расшифровать голосовое. Попробуй ещё раз или ответь текстом."
            )
        finally:
            try:
                await processing.delete()
            except TelegramBadRequest:
                pass

    return router
